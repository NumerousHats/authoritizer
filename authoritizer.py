import sys
from PyQt4 import QtCore, QtGui
from mainwindow import Ui_MainWindow
from selectcolumn import Ui_SelectcolsDialog
from selectsheet import Ui_SelectsheetDialog
from rundialog import Ui_Dialog
from preferencesdialog import Ui_PreferencesDialog
from runmodal import Ui_RunModal

import re
import os
import jellyfish
import unicodecsv as csv
import openpyxl

def cleanupImport(data, has_header):
    #data = [str(i) for i in data if i] # need something like this to cast numbers into strings? but it breaks unicode!
    data = [i for i in data if i != ""]
    if has_header:
        data = data[1:]
    data = list(set(data))
    return(data)


class StartQT4(QtGui.QMainWindow):
    def __init__(self, parent=None):
        QtGui.QWidget.__init__(self, parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.have_auth = False
        self.have_mess = False

        ###############
        ##### set up signals/slots
        ###############

        ##### menu items

        # use the trick from http://eli.thegreenplace.net/2011/04/25/passing-extra-arguments-to-pyqt-slot
        # to use same callback for two menu items
        self.ui.actionImport_auth.triggered.connect(lambda: self.importData("auth"))
        self.ui.actionImport_messy.triggered.connect(lambda: self.importData("messy"))
        self.ui.actionRun_matching.triggered.connect(self.runMatching)
        self.ui.actionExport_CSV.triggered.connect(self.exportCSV)
        self.ui.actionPreferences.triggered.connect(self.setPreferences)
        self.ui.actionQuit.triggered.connect(QtCore.QCoreApplication.instance().quit)

        ##### GUI elements

        self.ui.match_table.currentCellChanged.connect(self.updateTopHits)
        self.ui.tophit_list.itemDoubleClicked.connect(self.clickAssign)
        self.ui.createAuthority_button.clicked.connect(self.createAuth)
        self.ui.deleteAuthority_button.clicked.connect(self.deleteMatch)

        ###############
        ##### default preferences
        ###############

        self.cutoffs = {"lev": 10, "damlev": 10, "jaro": 0.6, "jarowink": 0.6, "mrac": 9999}
        self.display_similarity = False


    def importData(self, data_type):
        if data_type == "auth":
            window_name = "Import authorities"
        elif data_type == "messy":
            window_name = "Import nonstandard terms"
        
        fname = str(QtGui.QFileDialog.getOpenFileName(self, window_name, "~"))
        filename, file_extension = os.path.splitext(fname)

        if file_extension == ".csv":
            csv_fileh = open(fname, 'rU')
            try:
                dialect = csv.Sniffer().sniff(csv_fileh.read(1024))
                csv_fileh.seek(0)
                reader = csv.reader(csv_fileh, dialect=dialect)
                self.sample = [ reader.next() for i in range(20) ]
            except csv.Error:
                QtGui.QMessageBox.warning(self, 'Warning', 'File does not appear to be valid CSV')
                return

            dlg = StartSelectColumns(self, data_type)
            if dlg.exec_(): 
                selectcolumn_output = dlg.getValues()
            else:
                return

            # read the data from the selected column. reopen file for safety (even though it's absurdly inefficient)

            csv_fileh.close()
            csv_fileh = open(fname, 'rU')
            reader = csv.reader(csv_fileh, dialect=dialect)

            data = list()
            for row in reader:
                data.append(row[selectcolumn_output["column"]])
            data = cleanupImport(data, selectcolumn_output["header"])


        elif file_extension == ".txt":
            QtGui.QMessageBox.information(self, 'Information', 'Flat text import not yet supported.')
            return


        elif file_extension == ".xlsx":
            try:
                wb = openpyxl.load_workbook(fname)
                sheets = wb.get_sheet_names()
            except:
                QtGui.QMessageBox.warning(self, 'Warning', 'File does not appear to be valid Excel spreadsheet')
                return

            dlg = StartSelectSheet(sheets)
            if dlg.exec_(): 
                selected_sheet = dlg.getValues()
            else:
                return

            sheet = wb.get_sheet_by_name(selected_sheet)
            maxcol = sheet.get_highest_column()
            self.sample = [ [str(sheet.cell(row=i, column=j).value) for j in range(maxcol)] for i in range(20)]

            dlg = StartSelectColumns(self, data_type)
            if dlg.exec_(): 
                selectcolumn_output = dlg.getValues()
            else:
                return

            data = [ sheet.cell(row=i, column=selectcolumn_output["column"]).value for i in range(sheet.get_highest_row())]
            data = cleanupImport(data, selectcolumn_output["header"])

        elif file_extension == ".xls":
            QtGui.QMessageBox.information(self, 'Information', 'Excel .xls import not yet supported')
            return
        else:
            QtGui.QMessageBox.warning(self, 'Warning', 'File type {} is not supported'.format(file_extension))
            return

        if data_type == "auth":
            self.authorities = data
            self.have_auth = True
 
        elif data_type == "messy":
            self.mess = data
            self.have_mess = True

        else:
            QtGui.QMessageBox.critical(self, 'Error', 'Internal error: importData received unexpected argument')

        if self.have_auth and self.have_mess:
            self.ui.actionRun_matching.setEnabled(True)
 
    def runMatching(self):
        dlg = StartRunDialog() 
        if dlg.exec_(): 
            match_method = dlg.getValues() 
        else:
            return

        if match_method == "lev":
            match_function = jellyfish.levenshtein_distance
        elif match_method == "damlev":
            match_function = jellyfish.damerau_levenshtein_distance
        elif match_method == "jaro":
            match_function = jellyfish.jaro_distance
        elif match_method == "jarowink":
            match_function = jellyfish.jaro_winkler
        elif match_method == "mrac":
            match_function = jellyfish.match_rating_comparison
        else:
            QtGui.QMessageBox.critical(self, 'Warning', 'Internal error: runMatching received unexpected argument')

        dlg = StartRunModal(self, match_function, match_method) 
        if dlg.exec_(): 
            self.all_scores, self.matched_authorities = dlg.getValues() 
            self.ui.match_table.setRowCount(len(self.mess))
            self.ui.match_table.clearContents()
            self.updateTable()
            self.ui.actionExport_CSV.setEnabled(True)
        else:
            return



    def exportCSV(self):
        fname = QtGui.QFileDialog.getSaveFileNameAndFilter(self, 'Export CSV', '~', "*.csv")

        with open(fname[0], 'wb') as csvfile:
            csvwriter = csv.writer(csvfile)
            csvwriter.writerow(["Messy term", "Authority term"])

            for i in range(len(self.mess)):
                if self.matched_authorities[i]:
                    csvwriter.writerow([self.mess[i], self.matched_authorities[i]])

    def setPreferences(self):
        dlg = StartPreferences(self) 
        if dlg.exec_(): 
            preferences = dlg.getValues() 
        else:
            return

        self.display_similarity = preferences["display_similarity"]
        self.cutoffs = preferences["cutoffs"]

    def updateTable(self):
        for row in range(len(self.mess)):
            self.ui.match_table.setItem(row, 0, QtGui.QTableWidgetItem(self.mess[row]))
            if self.matched_authorities[row]:
                self.ui.match_table.setItem(row, 1, QtGui.QTableWidgetItem(self.matched_authorities[row]))

    def updateTopHits(self, row, column, oldrow, oldcolumn):
        self.ui.tophit_list.clear()
        for i in range(10):
            text = self.all_scores[row][i][0]
            if self.display_similarity:
                text = "{} ({:.3})".format(text, float(self.all_scores[row][i][1]))
            item = QtGui.QListWidgetItem(text)
            self.ui.tophit_list.addItem(item)

        if row != -1: # row gets set to -1 after deleteMatch(): ignore it and keep the old current_row
            self.current_row = row
        
    def clickAssign(self, item):
        # strip off trailing score if appended by updateTopHits
        authority = str(item.text())
        if self.display_similarity:
            self.matched_authorities[self.current_row] = re.sub(r'\s\([0-9.]+\)$', '', authority)
        else:
            self.matched_authorities[self.current_row] = authority
        self.updateTable()

    def createAuth(self):
        self.matched_authorities[self.current_row] = self.ui.new_authority.text()
        self.updateTable()

    def deleteMatch(self):
        self.matched_authorities[self.current_row] = False

        # there doesn't seem to be any way to clear a single cell (?!?)
        # so clear the entire table before re-rendering it
        self.ui.match_table.clearContents()
        self.updateTable()
        self.ui.match_table.setCurrentCell(self.current_row, 0)


class StartSelectColumns(QtGui.QDialog, Ui_SelectcolsDialog):
    def __init__(self, parent, type):
        QtGui.QDialog.__init__(self, parent)
        self.setupUi(self)

        nrow = len(parent.sample)
        ncol = len(parent.sample[0])

        if type == "auth":
            self.label.setText("Select column containing authorized terms")
        elif type == "messy":
            self.label.setText("Select column containing nonstandard terms")

        

        self.tableWidget.setColumnCount(ncol)
        self.tableWidget.setRowCount(nrow)
        self.tableWidget.cellClicked.connect(self.columnClicked)

        for row in range(nrow):
            for column in range(ncol):
                if parent.sample[row][column]:
                    self.tableWidget.setItem(row, column, QtGui.QTableWidgetItem(parent.sample[row][column]))

        self.tableWidget.setCurrentCell(0, 0)
        self.current_column = 0

    def columnClicked(self, row, column):
        self.current_column = column

    def getValues(self):
        return {"column": self.current_column, "header": self.checkBox.isChecked()}


class StartSelectSheet(QtGui.QDialog, Ui_SelectsheetDialog):
    def __init__(self, sheets=None, parent=None):
        QtGui.QDialog.__init__(self, parent)
        self.setupUi(self)

        for sheet in sheets:
            item = QtGui.QListWidgetItem(sheet)
            self.sheet_list.addItem(item)

        self.sheet_list.setCurrentRow(0)

    def getValues(self):
        return self.sheet_list.currentItem().text()


class StartPreferences(QtGui.QDialog, Ui_PreferencesDialog):
    def __init__(self, parent=None):
        QtGui.QDialog.__init__(self, parent)
        self.setupUi(self)

        self.lev_cutoff.setValue(parent.cutoffs["lev"])
        self.damlev_cutoff.setValue(parent.cutoffs["damlev"])
        self.jaro_cutoff.setValue(parent.cutoffs["jaro"])
        self.jarowink_cutoff.setValue(parent.cutoffs["jarowink"])
        self.mrac_cutoff.setValue(parent.cutoffs["mrac"])

        self.display_similarity.setChecked(parent.display_similarity)

    def getValues(self):
        cutoffs = {"lev": self.lev_cutoff.value(), 
                    "damlev": self.damlev_cutoff.value(),
                    "jaro": self.jaro_cutoff.value(),
                    "jarowink": self.jarowink_cutoff.value(),
                    "mrac": self.mrac_cutoff.value()}
        prefs = {"display_similarity": self.display_similarity.isChecked(), "cutoffs": cutoffs}

        return prefs

class StartRunDialog(QtGui.QDialog, Ui_Dialog):
    def __init__(self, parent=None):
        QtGui.QDialog.__init__(self, parent)
        self.setupUi(self)

        self.lev_rb.toggled.connect(self.levToggled)
        self.damlev_rb.toggled.connect(self.damlevToggled)
        self.jaro_rb.toggled.connect(self.jaroToggled)
        self.jarowink_rb.toggled.connect(self.jarowinkToggled)
        self.mrac_rb.toggled.connect(self.mracToggled)

        # lev_rb is selected by default within the .ui file,
        # therefore the following line is guaranteed to generate
        # a "toggled" event, and thereby makes sure that the default
        # matching algorithm (jaro-winkler) is set even if the user
        # doesn't actually click on anything

        self.jarowink_rb.click()

    def levToggled(self, state):
        if state:
            self.distfun = "lev"
    def damlevToggled(self, state):
        if state:
            self.distfun = "damlev"
    def jaroToggled(self, state):
        if state:
            self.distfun = 'jaro'
    def jarowinkToggled(self, state):
        if state:
            self.distfun = "jarowink"
    def mracToggled(self, state):
        if state:
            self.distfun = "mrac"

    def getValues(self):
        return self.distfun


class StartRunModal(QtGui.QDialog, Ui_RunModal):
    def __init__(self, parent, match_function, match_method):
        QtGui.QDialog.__init__(self, parent)
        self.setupUi(self)

        self.all_scores = list()
        self.matched_authorities = list()

        for m in parent.mess: # ideally, we want a progress bar for this loop
            print m
            scores = [ [x, match_function(m, unicode(x))] for x in parent.authorities ]
            scores = sorted(scores, key=lambda score: -score[1])[0:10]
            self.all_scores.append(scores)
            cutoff = parent.cutoffs[match_method]
            if match_method == "lev" or match_method == "damlev":
                self.matched_authorities.append(scores[0][0] if scores[0][1] < cutoff else False)
            else:
                self.matched_authorities.append(scores[0][0] if scores[0][1] > cutoff else False)

        self.accept()


    def getValues(self):
        return (self.all_scores, self.matched_authorities)


if __name__ == "__main__":
    app = QtGui.QApplication(sys.argv)
    myapp = StartQT4()
    myapp.show()
    sys.exit(app.exec_())


# for future preference setting

# from sys import platform as _platform

# if _platform == "linux" or _platform == "linux2":
#    # linux
# elif _platform == "darwin":
#    # MAC OS X
# elif _platform == "win32":
#    # Windows

# windows prefs are in the Application Data folder for the user or for all users.
# mac prefs are in /Users/username/Library/Preferences
# linux prefs are in ~/.authoritizer

